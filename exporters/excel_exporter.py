import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """
    Export Facebook posts to Excel.
    """

    def __init__(self, output_dir="output"):

        self.output_dir = output_dir

        os.makedirs(self.output_dir, exist_ok=True)

    # -----------------------------------------------------

    def export(self, posts, filename="posts.xlsx"):

        filepath = os.path.join(
            self.output_dir,
            filename
        )

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Facebook Posts"

        # ----------------------------------------------------
        # Headers
        # ----------------------------------------------------

        headers = [

            "Author",

            "Timestamp",

            "Post URL",

            "Post Text",

            "Reactions",

            "Comments",

            "Shares",

            "Like",

            "Love",

            "Care",

            "Haha",

            "Wow",

            "Sad",

            "Angry",

            "Images",

            "Videos",

            "Scraped At",

            "Source Page"

        ]

        for col, header in enumerate(headers, start=1):

            cell = worksheet.cell(row=1, column=col)

            cell.value = header

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ----------------------------------------------------
        # Data
        # ----------------------------------------------------

        row = 2

        for post in posts:

            worksheet.cell(row=row, column=1).value = post.author

            worksheet.cell(row=row, column=2).value = post.timestamp

            url_cell = worksheet.cell(row=row, column=3)

            url_cell.value = post.url

            if post.url:
                url_cell.hyperlink = post.url
                url_cell.style = "Hyperlink"

            text_cell = worksheet.cell(row=row, column=4)

            text_cell.value = post.text

            text_cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            worksheet.cell(row=row, column=5).value = post.reactions

            worksheet.cell(row=row, column=6).value = post.comments

            worksheet.cell(row=row, column=7).value = post.shares

            worksheet.cell(row=row, column=8).value = post.like

            worksheet.cell(row=row, column=9).value = post.love

            worksheet.cell(row=row, column=10).value = post.care

            worksheet.cell(row=row, column=11).value = post.haha

            worksheet.cell(row=row, column=12).value = post.wow

            worksheet.cell(row=row, column=13).value = post.sad

            worksheet.cell(row=row, column=14).value = post.angry

            image_cell = worksheet.cell(row=row, column=15)

            image_cell.value = "\n".join(post.images)

            image_cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            video_cell = worksheet.cell(row=row, column=16)

            video_cell.value = "\n".join(post.videos)

            video_cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

            worksheet.cell(row=row, column=17).value = post.scraped_at

            worksheet.cell(row=row, column=18).value = post.source_page

            row += 1

        # ----------------------------------------------------
        # Freeze Header
        # ----------------------------------------------------

        worksheet.freeze_panes = "A2"

        # ----------------------------------------------------
        # Auto Filter
        # ----------------------------------------------------

        worksheet.auto_filter.ref = worksheet.dimensions

        # ----------------------------------------------------
        # Auto Width
        # ----------------------------------------------------

        for column_cells in worksheet.columns:

            length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    if cell.value:

                        length = max(
                            length,
                            len(str(cell.value))
                        )

                except Exception:

                    pass

            if column == "D":

                worksheet.column_dimensions[column].width = 70

            elif column in ["O", "P"]:

                worksheet.column_dimensions[column].width = 50

            elif column == "C":

                worksheet.column_dimensions[column].width = 55

            elif column == "R":

                worksheet.column_dimensions[column].width = 45

            else:

                worksheet.column_dimensions[column].width = min(
                    max(length + 2, 18),
                    35
                )

        # ----------------------------------------------------
        # Row Height
        # ----------------------------------------------------

        for row_cells in worksheet.iter_rows(min_row=2):

            worksheet.row_dimensions[
                row_cells[0].row
            ].height = 120

        # ----------------------------------------------------
        # Save
        # ----------------------------------------------------

        workbook.save(filepath)

        print("\n==============================")
        print("EXCEL EXPORT")
        print("==============================")
        print(f"Posts Exported : {len(posts)}")
        print(f"Saved To       : {filepath}")
        print("==============================")